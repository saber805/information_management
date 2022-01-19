import json
from pip._vendor.distlib.compat import raw_input
import openpyxl

list_employee = []
filename = 'employee_date.json'


def show_menu():
    print("---" * 15)
    print('欢迎使用')
    print(' ')
    print("1.新建员工")
    print('2.显示所有员工')
    print('3.查找员工')
    print('4.删除员工')
    print('5.修改员工信息')
    print('6.导出为excel表格')
    print(' ')
    print('9.查看文件')
    print('0.退出系统')
    print("---" * 15)


def new_employee():
    name = input('员工姓名：')
    gender = input('性别（男/女）：')
    while len(gender) != 1:
        print('输入有误')
        gender = input('性别（男/女）：')
    phone = input('请输入联系方式：')
    while len(phone) != 11:
        phone = input('格式有误，请输入联系方式：')
    address = input('请输入家庭住址：')
    employee = {
        'name': name,
        'gender': gender,
        'phone': phone,
        'address': address
    }
    with open(filename) as f:
        list_employee = json.load(f)
        list_employee.append(employee)
        f.close()
    with open(filename, 'w') as f:
        json.dump(list_employee, f, indent=4)
        print('Add success {}'.format(employee['name']))
        print('Press enter to continue')
        input()
        f.close()


def show_all():         #there is a pause
   show_all2()
   print('Press enter to continue')
   input()


def search_employee():
    print('---'*15)
    print('查找员工')
    print('1,按姓名查找')
    print('2,按性别查找')
    print('3,按联系方式查找')
    print('4,按住址查找')
    number = input('选择功能（输入数字）：')
    while True:
        if number == '1':
            find_name = input('请输入要查找的姓名：')
            pri_()
            with open(filename) as f:
                list_employee = json.load(f)
                for employee in list_employee:
                    if employee['name'] == find_name:
                        print('{}\t\t{}\t\t{}\t\t{}\t\t'.format(employee['name'], employee['gender'],
                                                            employee['phone'], employee['address']))
                f.close()
            print('Press enter to continue')
            input()
            return
        elif number == '2':
            find_gender = input('输入性别:')
            while len(find_gender) != 1:
                find_gender = input('输入有误，重新输入：')
            pri_()
            with open(filename) as f:
                list_employee = json.load(f)
                for employee in list_employee:
                    if employee['gender'] == find_gender:
                        print('{}\t\t{}\t\t{}\t\t{}\t\t'.format(employee['name'], employee['gender'],
                                                                employee['phone'], employee['address']))
                f.close()
            print('Press enter to continue')
            input()
            return
        elif number == '3':
            find_number = input('输入要查找的联系方式：')
            pri_()
            with open(filename) as f:
                list_employee = json.load(f)
                for employee in list_employee:
                    if employee['phone'] == find_number:
                        print('{}\t\t{}\t\t{}\t\t{}\t\t'.format(employee['name'], employee['gender'],
                                                                employee['phone'], employee['address']))
                f.close()
            print('Press enter to continue')
            input()
            return
        elif number == '4':
            find_address = input('输入要查找的地址：')
            pri_()
            with open(filename) as f:
                list_employee = json.load(f)
                for employee in list_employee:
                    if employee['address'] == find_address:
                        print('{}\t\t{}\t\t{}\t\t{}\t\t'.format(employee['name'], employee['gender'],
                                                                employee['phone'], employee['address']))
                f.close()
            print('Press enter to continue')
            input()
            return
        else:
            print('输入有误，请重新输入')


def del_employee():
    show_all2()
    find_name = input('请输入要删除员工的姓名：')
    pri_()
    with open(filename) as f:
        list_employee = json.load(f)
        for employee in list_employee:
            if employee['name'] == find_name:
                list_employee.remove(employee)
        f.close()
    with open(filename, 'w') as f:
        json.dump(list_employee, f, indent=4)
        f.close()


def modify_employee():
    global dict_employee
    dict_employee = {}
    show_all2()
    find_sb = input('请输入要修改的员工姓名：')
    with open(filename) as f:
        list_employee = json.load(f)
        for employee in list_employee:      #employee是个字典
            if employee['name'] == find_sb:
                dict_employee = employee
                print('{}\t\t{}\t\t{}\t\t{}\t\t'.format(employee['name'], employee['gender'],
                                                        employee['phone'], employee['address']))
        if not any(dict_employee):
            print("NOT FOUND!!!")
            input()
            return
        f.close()
    print('')
    print('---' * 15)
    print('要修改\t\t1，姓名\t\t2,联系方式\t\t3,家庭住址\t\t4，性别\t\t')
    while True:
        str = raw_input("please input the number:")
        if str.isdigit():   #接收全为数字为True
            if str == '1':
                print('原名：{}'.format(dict_employee['name']))
                new_name = input('输入新姓名：')
                displace('name', new_name)
                break
            elif str == '2':
                print('原联系方式：{}'.format(dict_employee['phone']))
                new_phone = input('新联系方式：')
                displace('phone', new_phone)
                break
            elif str == '3':
                print('原住址：{}'.format(dict_employee['address']))
                new_address = input('新住址：')
                displace('address', new_address)
                break
            elif str == '4':
                print('{}：{}'.format(dict_employee['name'], dict_employee['gender']))
                y = input('确认修改？（y/n）：')
                if y == 'y':
                    if dict_employee['gender'] == '男':
                        new_gender = '女'
                    elif dict_employee['gender'] == '女':
                        new_gender = '男'
                    displace('gender', new_gender)
                elif y == 'n':
                    break
                break
        elif str == 'q':
            break
        else:
            print('搁着找bug呢？')
            input()
            print('重新输入，别整些没用的（q.退出）')


def show_all2():           #no pause
    with open(filename, 'r') as f:
        data = json.load(f)
        if len(data) == 0:
            print('No employee ')
            print('Press enter to continue')
            input()
            return
        pri_()
        for a in data:
            print('{}\t\t{}\t\t{}\t\t{}\t\t'.format(a['name'], a['gender'],
                                                    a['phone'], a['address']))
        f.close()


def pri_():
    for name in ['姓名', '性别', '联系方式', '家庭住址']:
        print(name, end='\t\t')
    print('')
    print('---' * 15)


def export_to_excel():
    # 定义要写入的行和列的值
    with open(filename) as f:
        list_employee = json.load(f)
            # 定义excel的sheet_name  "xlsx格式测试表 "
        f.close()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet_name = "xlsx格式表"
    sheet.title = sheet_name

    path = 'C:/Users/acer/Desktop/xlsx格式demo.xlsx'
    print('默认路径：{}'.format(path))
    print('是否修改：（y/n）')
    a = input()
    if a == 'y':
        path = input('新路径：')
    temporary_val_k = ['姓名', '性别', '联系方式', '家庭住址']
    for a in range(0, 4):
        sheet.cell(row=1, column=a+1, value=temporary_val_k[a])     #打印表头
    i = 0
    while i < len(list_employee):    #0~len(list)   i：0~4
        temporary_val = []
        for val in list_employee[i].values():     #list_employee is a list ,每个值为字典
            temporary_val.append(val)   #此时temporary_val 是列表，有四个值
        for j in range(len(temporary_val)):
            sheet.cell(i+2, j+1, temporary_val[j])
        i = i+1
    workbook.save(path)
    print("xlsx格式表格写入数据成功!")
    print('Press enter to continue')
    input()


def displace(a, b):
    with open(filename) as f:
        list_employee = json.load(f)
        for employee in list_employee:  # 在json中寻找与原名重复的员工
            if employee == dict_employee:
                list_employee.remove(employee)
                dict_employee[a] = b  # employee要覆盖原信息
                list_employee.append(dict_employee)
                f.close()
    with open(filename, 'w') as f:
        json.dump(list_employee, f, indent=4)
        print('modified successfully')
        input()
        f.close()


def read():
    with open(filename) as f:
        data = json.load(f)
        print(data)
        input()
        f.close()


def new_employee2(number):
    employee = {
        'name': '杰洛特',
        'gender': '男',
        'phone': number,
        'address': 'lllllssss'
    }
    with open(filename) as f:
        list_employee = json.load(f)
        list_employee.append(employee)
        f.close()
    with open(filename, 'w') as f:
        json.dump(list_employee, f, indent=4)
        print('Add success {}'.format(employee['name']))
        f.close()


